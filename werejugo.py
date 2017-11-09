from Registry import Registry
import requests
import json
import struct
import openpyxl
import glob
import sys
import argparse
import datetime
import re
import lxml.etree
import Evtx.Evtx
import Evtx.Views
import Evtx.Nodes
import os

try:
    input=raw_input
except:
    pass

def xml_records(filename):
    #Code based on https://raw.githubusercontent.com/williballenthin/python-evtx/master/scripts/evtx_filter_records
    """
    If the second return value is not None, then it is an
      Exception encountered during parsing.  The first return value
      will be the XML string.

    @type filename str
    @rtype: generator of (etree.Element or str), (None or Exception)
    """
    with Evtx.Evtx.Evtx(filename) as evtx:
        for xml, record in Evtx.Views.evtx_file_xml_view(evtx.get_file_header()):
            try:
                yield lxml.etree.fromstring(b"<?xml version=\"1.0\" standalone=\"yes\" ?>%s" % xml), None
            except lxml.etree.XMLSyntaxError as e:
                yield xml, e

def get_child(node, tag, ns="{http://schemas.microsoft.com/win/2004/08/events/event}"):
    #Code based on https://raw.githubusercontent.com/williballenthin/python-evtx/master/scripts/evtx_filter_records
    """
    @type node: etree.Element
    @type tag: str
    @type ns: str
    """
    return node.find("%s%s" % (ns, tag))

def google_cheat(networks):
    #cheats using a google api key by looking up the data directly with the browser location service
    url = 'https://maps.googleapis.com/maps/api/browserlocation/json?browser=firefox&sensor=true'
    for eachnetwork in networks:
        url += '&wifi=mac:{0}|ssid:{2}|ss:{1}'.format(*eachnetwork)
    webresp = requests.get(url)
    if webresp.status_code == 200:
        x = webresp.json()
    else:
        print("URL: {0}".format(url))
        print(webresp.status_code, webresp.reason, webresp.content)
        return("Not Found", "Not Found")
    accuracy =  "Accuracy of %s square meters" % (x['accuracy'])
    geourl =  "http://maps.google.com/maps?q=%.9f,%.9f&z=15" % (x['location']['lat'], x['location']['lng'])
    return (accuracy, geourl)

def parse_6100(path_to_evtx):
    results = []
    for node, err in xml_records(path_to_evtx):
        if err is not None:
            continue
        EventMetadata = get_child(node, "System")
        EventData = get_child(node, "EventData")
        if 6100 == int(get_child(EventMetadata, "EventID").text):
            SystemTime = get_child(EventMetadata, "TimeCreated").attrib
            #import pdb;pdb.set_trace()
            EventDescription = "".join([x.text for x in EventData])
            sids = re.findall(r"(\w\w-\w\w-\w\w-\w\w-\w\w-\w\w).*?Infra.*?&lt;\s+(-\d{1,3})\s+\d{1,8}\s+([()\w]+)",EventDescription)
            if not sids:
                continue
            accuracy,url = google_cheat(sids)
            results.append((str(SystemTime), EventDescription, accuracy, url))
    return results

def parse_wlan_autoconfig(path_to_evtx):
    results = []
    Connect_Events = [8001 ] 
    Disconnect_Events = [8003] 
    Failed_Events = [8002, 11002]
    Still_Connected  = [11004, 11005, 11010, 11006, 12011, 12012, 12013]
    for node, err in xml_records(path_to_evtx):
        if err is not None:
            continue
        EventMetadata = get_child(node, "System")
        EventData = get_child(node, "EventData")
        SystemTime = get_child(EventMetadata, "TimeCreated").attrib
        EventDescription = "\n".join(["{0}:{1}".format(x.values()[0],str(x.text)) for x in EventData])
        EventID = int(get_child(EventMetadata, "EventID").text)
        if EventID in Connect_Events:
            eventtype = "Connect"
        elif EventID in Disconnect_Events:
            eventtype = "Disconnect"
        elif EventID in Failed_Events:
            eventtype = "Failed Connection"
        else:
            continue
        ProfileName = re.search(r"SSID:(\S+)", EventDescription).group(1)
        BSSID = profile_table.get(ProfileName, "MAC-UNKNOWN")
        mapurl = "Unable to determine MAC for SSID"
        if BSSID != "MAC-UNKNOWN":
            mapurl = "Unable to find MAC on Wigle"
            result = wigle_search(netid=BSSID)
            if len(result['results']):
                for eachresult in result['results']:
                    mapurl =  "http://maps.google.com/maps?q=%.9f,%.9f&z=15" % (eachresult['trilat'], eachresult['trilong'])
        results.append((eventtype, BSSID, ProfileName, SystemTime['SystemTime'], mapurl ))
    return results

def get_reg_value(path_to_file, path, key):
    rh = Registry.Registry(path_to_file)
    rk = rh.open(path)
    return rk.value(key).value()

wigle_cache={}
def wigle_search(**kwargs):
    #Lookup like this  wigle_search(netid="ff:ff:ff:ff:ff")
    if kwargs['netid'] in wigle_cache:
        print("From Cache")
        return wigle_cache.get(kwargs['netid'])
    url = "https://api.wigle.net/api/v2/network/search"
    if not options.WIGLE_USER or not options.WIGLE_PASS:
            return {'results':[{'trilat':0, 'trilong': 0}]}
    try:
        webresp = requests.get(url, auth = (options.WIGLE_USER,options.WIGLE_PASS), params = kwargs)
        if webresp.status_code != 200:
            print("*"*25,"There was an error from Wigle. {0}".format(webresp.reason))
        result = webresp.json()
        wigle_cache[kwargs['netid']] = result
        if result['success']==False:
            raise(Exception("*"*25, "Wigle Request Failed {0}".format(result['message'])))
    except Exception as e:
        cont = input("Bad things happened while talking to Wigle. {0}, {1}  Continue? [Y/N] ".format(webresp.reason, str(e)))
        if cont.lower()=="n":
            raise(Exception("Wigle connection error. {0} {1} {2}".format(webresp.status_code,webresp.reason,str(e))))
    return result

def reg_date(dateblob):
    weekday = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
    year,month,day,date,hour,minute,sec,micro = struct.unpack('<8H', dateblob)
    dt = "%s, %02d/%02d/%04d %02d:%02d:%02d.%s" % (weekday[day],month,date,year,hour,minute,sec,micro)
    dtp = datetime.datetime.strptime(dt,"%A, %m/%d/%Y %H:%M:%S.%f")
    return dtp

def get_profile_info(reg_handle, ProfileGuid):
    nametypes = {'47':"Wireless", "06":"Wired", "17":"Broadband"}
    guid = b"Microsoft\Windows NT\CurrentVersion\NetworkList\Profiles\%s" % (ProfileGuid)
    regkey = reg_handle.open(guid)
    NameType ="%02x"  % regkey.value("NameType").value()
    nettype = nametypes.get(str(NameType), "Unknown Type"+str(NameType))
    FirstConnect = reg_date(regkey.value("DateCreated").value())
    LastConnect  = reg_date(regkey.value("DateLastConnected").value())
    return nettype, FirstConnect, LastConnect

def network_history(reg_handle):
    global profile_table
    reg_results = []
    for eachsubkey in reg_handle.open(b"Microsoft\Windows NT\CurrentVersion\NetworkList\Signatures\Unmanaged").subkeys():
        DefaultGatewayMac = eachsubkey.value("DefaultGatewayMac").value()
        BSSID = ':'.join(DefaultGatewayMac.encode("HEX")[i:i+2] for i in range(0,12,2))
        Description = eachsubkey.value("Description").value()
        DnsSuffix = eachsubkey.value("DnsSuffix").value()
        SSID = eachsubkey.value("FirstNetwork").value()
        ProfileGuid = eachsubkey.value("ProfileGuid").value()
        nettype,first,last = get_profile_info(reg_handle, ProfileGuid)
        mapurl="Not found"
        #print "TYPE:",nettype, " SSID:",SSID, "Description:", Description, "DNS SUFFIX", DnsSuffix, "MAC ADDRESS:", BSSID
        if nettype=="Wireless":
            result = wigle_search(netid=BSSID)
            if len(result['results']):
                for eachresult in result['results']:
                    #print "Found: ", SSID, BSSID
                    #print "First connected: ",first
                    #print "Last Connected : ",last  
                    mapurl =  "http://maps.google.com/maps?q=%.9f,%.9f&z=15" % (eachresult['trilat'], eachresult['trilong'])
            #print google_cheat(BSSID, SSID)
            profile_table[SSID] = BSSID
        reg_results.append((BSSID,SSID,Description,DnsSuffix,nettype,first,last,mapurl))
    return reg_results

def reg_basic_info():
    #Retreive basic computer data
    print("Enumerating Basic Computer Data")
    osversion = get_reg_value("SOFTWARE", "Microsoft\Windows NT\CurrentVersion", "CurrentVersion")
    which_control_set = get_reg_value("SYSTEM","Select","Current")
    ControlSet = "ControlSet%03d" % (which_control_set)
    comp_name = get_reg_value("SYSTEM",ControlSet+"\Control\ComputerName\ComputerName","ComputerName")
    time_zone = get_reg_value("SYSTEM",ControlSet+"\Control\TimeZoneInformation","DayLightName")
    print(osversion, which_control_set, comp_name, time_zone)

def reg_interface_info():
    #Retrieve local area network connections
    print("Enumerating Local Network Interface Data")
    rh = Registry.Registry("SYSTEM")
    interfaces = rh.open(ControlSet+"\Services\Tcpip\Parameters\Interfaces").subkeys()
    for eachinterface in interfaces:
        try:
            print(eachinterface.value("DhcpIPAddress").value())
            print(eachinterface.value("DhcpDomain").value())
        except:
            pass

parser = argparse.ArgumentParser(description="Given a directory of SOFTWARE registry hives it will enumerate all the wireless.")
parser.add_argument("--IMAGE_MOUNT_POINT","-i",default='', help ="Geolocate all sources searching the SYSTEM volume of this mounted Windows Images . example: z:\image\C:")
#parser.add_argument("--recurse", help = "Recursively, Exhaustively search the entire drive for any usable file instead of looking in default locatoins.", action="store_true", )
parser.add_argument("--SOFTWARE_REGISTRY","-r", default="", help ="Location of the SOFTWARE registry keys. (Supports Wildcards) example: c:\windows\system32\config\SYSTEM*")
parser.add_argument("--SYSTEM_EVENTS","-e", default="", help ="Location of System Event Log specified. (Support Wildcards) Example: c:\windows\system32\Winevt\Logs\System.evtx)")
parser.add_argument("--WLAN_EVENTS","-w", default="", help ="Location of WLAN Autoconfig Event Log specified. (Support Wildcards) Example: c:\windows\system32\Winevt\Logs\Microsoft-Windows-WLAN-AutoConfig perational.evtx)")
parser.add_argument("--OUTPUT","-o", default="werejugo_output.xlsx", help ="Path and filename of the XLSX file to create.")
parser.add_argument("--WIGLE_USER", "-u", help="Wigle.net API Username.  Required for geolocating  See https://wigle.net/account")
parser.add_argument("--WIGLE_PASS", "-p", help="Wigle.net API Password.  Required for geolocating  See https://wigle.net/account")

options = parser.parse_args()

#if options.recurse:
#    print("Exhaustive search not implemented yet.")

reg_files = os.path.join(options.IMAGE_MOUNT_POINT,options.SOFTWARE_REGISTRY) 
evt_files = os.path.join(options.IMAGE_MOUNT_POINT,options.SYSTEM_EVENTS)
wlan_evts_files = os.path.join(options.IMAGE_MOUNT_POINT,options.WLAN_EVENTS)
if "*" in reg_files:
    reg_files = glob.glob(reg_files)
else:
    reg_files = [reg_files]
if "*" in evt_files:
    evt_files = glob.glob(evt_files)
else:
    evt_files = [evt_files]
if "*" in wlan_evts_files:
    wlan_evts_files = glob.glob(wlan_evts_files)
else:
    wlan_evts_files = [wlan_evts_files]


if not options.WIGLE_USER or not options.WIGLE_PASS:
        print("\nA Wigle API username and password is required for geolocation resolution.  Obtain one https://wigle.net/account.")
        print("All Locations will be 0,0.\n\n")

#Retrieve Wifi history
print("Enumerating Wireless History and GPS Locations")
target_wb = openpyxl.Workbook()

#Create a Consolidated Timeline tab in the spreadsheet
timeline_sheet = target_wb.create_sheet(title="Consolidated Timeline")
columns = ['Date','BSSID', 'SSID','Geolocation URL']
for column, value in enumerate(columns):
    timeline_sheet.cell(row = 1, column = column+1).value = value


#Do the wireless Data sheet
if options.SOFTWARE_REGISTRY:
    print("Enumerating Wireless Profiles in the Registry...")
    xls_sheet = target_wb.create_sheet(title="Wireless Data")
    columns = ['BSSID', 'SSID','Description', 'DNSSufix', 'Type', 'First Seen', 'Last Seen', 'Geolocation URL']
    profile_table = {}
    for column, value in enumerate(columns):
        xls_sheet.cell(row = 1, column = column+1).value = value
    for eachfile in reg_files:
        try:
            reg_handle = Registry.Registry(eachfile)
            reg_handle.open(b"Microsoft\Windows NT\CurrentVersion\NetworkList\Signatures\Unmanaged")
        except Exception as e:
            print("Unable to process file {0} : {1} ".format(eachfile,e))
        else:
            reg_data = network_history(reg_handle)
            for row, row_data in enumerate(reg_data):
                for column, value in enumerate(row_data):
                    xls_sheet.cell(row = row+2, column = column+1).value = value
                if row_data[7].startswith("http"):
                    timeline_sheet.append([row_data[5], row_data[0], row_data[1],row_data[7]])
                    timeline_sheet.append([row_data[6], row_data[0], row_data[1],row_data[7]])
        try:
            reg_handle = Registry.Registry(eachfile)
            reg_handle.open(b"Microsoft\Windows NT\CurrentVersion\NetworkList\Signatures\Managed")
        except Exception as e:
            print("Unable to process file {0} : {1} ".format(eachfile,e))
        else:
            reg_data = network_history(reg_handle)
            profile_table[reg_data[1]] = reg_data[0]
            for row, row_data in enumerate(reg_data):
                for column, value in enumerate(row_data):
                    xls_sheet.cell(row = row+2, column = column+1).value = value
                if row_data[7].startswith("http"):
                    timeline_sheet.append([row_data[5], row_data[0], row_data[1],row_data[7]])
                    timeline_sheet.append([row_data[6], row_data[0], row_data[1],row_data[7]])


#Do the Event 6100 sheet
if options.SYSTEM_EVENTS:
    print("Enumerating Event 6100 data from SYSTEM event logs...")
    print("Google disabled the required API.  This feature is temporarily unavailable.")
    """xls_sheet = target_wb.create_sheet(title="Event 6100 Data")
    columns = ['System Time', 'Data','Geolocation Accuracy', 'Geolocation URL']
    for column, value in enumerate(columns):
        xls_sheet.cell(row = 1, column = column+1).value = value
    for eachfile in evt_files:
        #try:
        evt_data  = parse_6100(eachfile)
        #except Exception as e:
        #    print("Unable to process file {0} : {1} ".format(eachfile,e))
        #    continue
        for row, row_data in enumerate(evt_data):
            for column, value in enumerate(row_data):
                xls_sheet.cell(row = row+2, column = column+1).value = value
            if row_data[3].startswith("http"):
                timeline_sheet.append([row_data[0], row_data[1], "hmm",row_data[3]])
    """

#Do WLAN_AUTOCONFIG
if options.WLAN_EVENTS:
    print("Enumerating WLAN Autoconfig Log data")
    xls_sheet = target_wb.create_sheet(title="WLAN Events Data")
    columns = ['ACTION', 'BSSID','SSID', 'System Time', "Geolocation URL"]
    for column, value in enumerate(columns):
        xls_sheet.cell(row = 1, column = column+1).value = value
    for eachfile in wlan_evts_files:
        try:
            evt_data  = parse_wlan_autoconfig(eachfile)
        except Exception as e:
            print("Unable to process file {0} : {1} ".format(eachfile,e))
            continue
        for row, row_data in enumerate(evt_data):
            for column, value in enumerate(row_data):
                xls_sheet.cell(row = row+2, column = column+1).value = value
            if row_data[4].startswith("http"):
                timeline_sheet.append([row_data[3], row_data[1], row_data[2],row_data[4]])


firstsheet=target_wb.get_sheet_by_name("Sheet")
target_wb.remove_sheet(firstsheet)
target_wb.save(options.OUTPUT)